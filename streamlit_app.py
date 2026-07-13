import datetime
from datetime import timedelta
import time, hmac, hashlib, base64, json
import urllib.request, urllib.parse, urllib.error
import pandas as pd
import altair as alt
import streamlit as st

st.set_page_config(page_title="Koras 광고 대시보드", layout="wide")

# ========================================================
#  디자인 (코라스 파랑 톤)
# ========================================================
BLUE = "#2563EB"
BLUE_SCHEME = ["#2563EB", "#7AA5F5", "#1E3A8A", "#9DBEF7"]
st.markdown("""
<style>
.block-container {padding-top: 3.5rem; padding-bottom: 2.5rem; max-width: 1200px;}
hr {margin: 1.1rem 0;}
#MainMenu, footer {visibility: hidden;}
.k-tag {display:inline-flex; align-items:center; gap:6px; font-size:11px; letter-spacing:0.08em;
        font-weight:500; color:#2563EB; background:rgba(37,99,235,0.10);
        padding:4px 10px; border-radius:999px; margin:4px 0 8px;}
.k-h1 {font-size:23px; font-weight:700; line-height:1.2; margin:0;}
.k-sec {display:flex; align-items:center; gap:8px; margin:0 0 10px;}
.k-bar {width:3px; height:15px; background:#2563EB; border-radius:2px; display:inline-block;}
.k-sec-t {font-size:15px; font-weight:600;}
.k-hero {border:0.5px solid rgba(37,99,235,0.25); border-radius:14px; padding:18px 20px;
         margin-bottom:6px; background:rgba(37,99,235,0.04);}
.k-cell {background:rgba(128,128,128,0.06); border-radius:10px; padding:14px 16px;}
.k-hero .lbl {font-size:12px; color:rgba(128,128,128,0.95); margin-bottom:6px;}
.k-hero .num {font-size:28px; font-weight:700; line-height:1.05; letter-spacing:-0.01em;}
.k-pill {display:inline-flex; align-items:center; gap:2px; font-size:11px; font-weight:600;
         padding:2px 8px; border-radius:999px; margin-top:9px;}
.k-up {background:rgba(22,163,74,0.14); color:#15803D;}
.k-dn {background:rgba(220,38,38,0.14); color:#DC2626;}
.k-strip {margin-top:14px; padding-top:12px; border-top:0.5px solid rgba(128,128,128,0.2);
          font-size:12px; color:rgba(128,128,128,0.95); display:flex; gap:20px; flex-wrap:wrap;}
.k-strip b {color:#2563EB; font-weight:700;}
table.k-tbl {width:100%; border-collapse:collapse; font-size:13px; color:inherit;}
table.k-tbl th {background:rgba(37,99,235,0.07); color:#2563EB; padding:9px 12px; font-weight:600;}
table.k-tbl td {padding:10px 12px; border-top:0.5px solid rgba(128,128,128,0.18);}
table.k-tbl tr.tot td {border-top:1.5px solid rgba(37,99,235,0.35);
          background:rgba(37,99,235,0.06); font-weight:700;}
.k-dot {width:7px; height:7px; border-radius:50%; display:inline-block; margin-right:7px; vertical-align:middle;}
.k-wrap {border:0.5px solid rgba(128,128,128,0.22); border-radius:14px; overflow:hidden;}
</style>
""", unsafe_allow_html=True)

START_DATE = st.secrets.get("GOOGLE_START_DATE", "2026-04-01")
TODAY = datetime.date.today().isoformat()


# ========================================================
#  SNS/DA 데이터 (구글 + 메타)
# ========================================================
@st.cache_data(ttl=3600)
def load_google():
    from google.ads.googleads.client import GoogleAdsClient
    cfg = {
        "developer_token": st.secrets["GOOGLE_DEVELOPER_TOKEN"],
        "client_id": st.secrets["GOOGLE_CLIENT_ID"],
        "client_secret": st.secrets["GOOGLE_CLIENT_SECRET"],
        "refresh_token": st.secrets["GOOGLE_REFRESH_TOKEN"],
        "use_proto_plus": True,
    }
    customer_id = str(st.secrets["GOOGLE_CUSTOMER_ID"])
    client = GoogleAdsClient.load_from_dict(cfg)
    ga = client.get_service("GoogleAdsService")
    data = {}
    camp_q = f"""
        SELECT segments.date, campaign.name, campaign.advertising_channel_type,
               metrics.impressions, metrics.clicks, metrics.cost_micros, metrics.conversions
        FROM campaign
        WHERE segments.date BETWEEN '{START_DATE}' AND '{TODAY}'
    """
    ch_map = {}
    for batch in ga.search_stream(customer_id=customer_id, query=camp_q):
        for r in batch.results:
            key = (str(r.segments.date), r.campaign.name)
            ch = str(r.campaign.advertising_channel_type)
            ch_map[r.campaign.name] = "google_sa" if "SEARCH" in ch.upper() else "google"
            data[key] = {"impressions": int(r.metrics.impressions),
                         "clicks": int(r.metrics.clicks), "views": 0,
                         "cost": r.metrics.cost_micros / 1_000_000,
                         "conversions": float(r.metrics.conversions)}
    view_q = f"""
        SELECT segments.date, campaign.name, metrics.video_trueview_views
        FROM campaign
        WHERE segments.date BETWEEN '{START_DATE}' AND '{TODAY}'
    """
    try:
        for batch in ga.search_stream(customer_id=customer_id, query=view_q):
            for r in batch.results:
                key = (str(r.segments.date), r.campaign.name)
                if key not in data:
                    data[key] = {"impressions": 0, "clicks": 0, "views": 0, "cost": 0.0, "conversions": 0.0}
                data[key]["views"] += int(r.metrics.video_trueview_views)
    except Exception:
        pass
    rows = [{"date": d, "platform": ch_map.get(c, "google"), "campaign": c, **m}
            for (d, c), m in data.items()]
    return pd.DataFrame(rows)


@st.cache_data(ttl=3600)
def load_meta():
    token = st.secrets.get("META_ACCESS_TOKEN")
    acct = st.secrets.get("META_AD_ACCOUNT_ID")
    if not token or not acct:
        return pd.DataFrame()
    try:
        from facebook_business.api import FacebookAdsApi
        from facebook_business.adobjects.adaccount import AdAccount
        FacebookAdsApi.init(access_token=token)
        account = AdAccount(f"act_{acct}")
        params = {"level": "adset", "time_range": {"since": START_DATE, "until": TODAY}, "time_increment": 1}
        fields = ["campaign_name", "adset_name", "impressions", "clicks", "spend", "reach", "actions", "date_start"]
        rows = []
        for row in account.get_insights(fields=fields, params=params):
            link_click = 0
            for a in (row.get("actions") or []):
                if a.get("action_type") == "link_click":
                    link_click = int(float(a.get("value", 0)))
                    break
            rows.append({"date": str(row.get("date_start")), "platform": "meta",
                         "campaign": row.get("adset_name") or row.get("campaign_name"),
                         "impressions": int(row.get("impressions", 0) or 0),
                         "clicks": int(row.get("clicks", 0) or 0),
                         "views": int(row.get("reach", 0) or 0),
                         "cost": float(row.get("spend", 0) or 0),
                         "conversions": float(link_click)})
        return pd.DataFrame(rows)
    except Exception as e:
        st.warning(f"메타 데이터를 불러오지 못했어요: {e}")
        return pd.DataFrame()


@st.cache_data(ttl=3600)
def load_sns():
    g = load_google()
    m = load_meta()
    df = pd.concat([g, m], ignore_index=True)
    if not df.empty:
        df["date"] = pd.to_datetime(df["date"])
    return df


# ========================================================
#  검색광고(SA) 데이터 (네이버)
# ========================================================
NV_BASE = "https://api.searchad.naver.com"


def _nv_sign(secret, ts, method, path):
    msg = f"{ts}.{method}.{path}"
    return base64.b64encode(hmac.new(secret.encode(), msg.encode(), hashlib.sha256).digest()).decode()


def _nv_get(path, query=None):
    cid = str(st.secrets["NAVER_CUSTOMER_ID"])
    key = st.secrets["NAVER_API_KEY"]
    secret = st.secrets["NAVER_SECRET_KEY"]
    ts = str(int(time.time() * 1000))
    url = NV_BASE + path
    if query:
        url += "?" + urllib.parse.urlencode(query)
    hdr = {"X-Timestamp": ts, "X-API-KEY": key, "X-Customer": cid,
           "X-Signature": _nv_sign(secret, ts, "GET", path)}
    req = urllib.request.Request(url, headers=hdr)
    with urllib.request.urlopen(req, timeout=60) as r:
        return json.loads(r.read().decode())


def _chunks(lst, n):
    for i in range(0, len(lst), n):
        yield lst[i:i + n]


def _naver_ready():
    return all(k in st.secrets for k in ("NAVER_CUSTOMER_ID", "NAVER_API_KEY", "NAVER_SECRET_KEY"))


@st.cache_data(ttl=3600)
def nv_summary(since, until):
    campaigns = _nv_get("/ncc/campaigns")
    ids = [c["nccCampaignId"] for c in campaigns]
    fields = ["impCnt", "clkCnt", "salesAmt", "ccnt"]
    tot = {"impressions": 0, "clicks": 0, "cost": 0.0, "conversions": 0.0}
    for batch in _chunks(ids, 100):
        q = {"ids": ",".join(batch), "fields": json.dumps(fields),
             "timeRange": json.dumps({"since": since, "until": until})}
        try:
            for row in _nv_get("/stats", q).get("data", []):
                tot["impressions"] += int(row.get("impCnt", 0) or 0)
                tot["clicks"] += int(row.get("clkCnt", 0) or 0)
                tot["cost"] += float(row.get("salesAmt", 0) or 0)
                tot["conversions"] += float(row.get("ccnt", 0) or 0)
        except urllib.error.HTTPError:
            pass
    return tot


@st.cache_data(ttl=3600)
def nv_daily(since, until):
    campaigns = _nv_get("/ncc/campaigns")
    ids = [c["nccCampaignId"] for c in campaigns]
    fields = ["impCnt", "clkCnt", "salesAmt"]
    d0 = datetime.date.fromisoformat(since)
    d1 = datetime.date.fromisoformat(until)
    rows = []
    cur = d0
    while cur <= d1:
        day = cur.isoformat()
        t = {"impressions": 0, "clicks": 0, "cost": 0.0}
        for batch in _chunks(ids, 100):
            q = {"ids": ",".join(batch), "fields": json.dumps(fields),
                 "timeRange": json.dumps({"since": day, "until": day})}
            try:
                for row in _nv_get("/stats", q).get("data", []):
                    t["impressions"] += int(row.get("impCnt", 0) or 0)
                    t["clicks"] += int(row.get("clkCnt", 0) or 0)
                    t["cost"] += float(row.get("salesAmt", 0) or 0)
            except urllib.error.HTTPError:
                pass
        rows.append({"date": day, **t})
        cur += timedelta(days=1)
    df = pd.DataFrame(rows)
    if not df.empty:
        df["date"] = pd.to_datetime(df["date"])
    return df


@st.cache_data(ttl=3600)
def nv_keywords(since, until):
    adgroups = _nv_get("/ncc/adgroups")
    all_kw = []
    for ag in adgroups:
        try:
            all_kw.extend(_nv_get("/ncc/keywords", {"nccAdgroupId": ag.get("nccAdgroupId")}))
        except Exception:
            pass
    ids = [k["nccKeywordId"] for k in all_kw]
    fields = ["impCnt", "clkCnt", "salesAmt", "cpc"]
    stat = {}
    for batch in _chunks(ids, 100):
        q = {"ids": ",".join(batch), "fields": json.dumps(fields),
             "timeRange": json.dumps({"since": since, "until": until})}
        try:
            for row in _nv_get("/stats", q).get("data", []):
                stat[row.get("id")] = row
        except urllib.error.HTTPError:
            pass
    rows = []
    for k in all_kw:
        s = stat.get(k.get("nccKeywordId"), {})
        rows.append({
            "키워드": k.get("keyword"),
            "노출수": int(s.get("impCnt", 0) or 0),
            "클릭수": int(s.get("clkCnt", 0) or 0),
            "현재 입찰가": int(k.get("bidAmt", 0) or 0),
            "클릭 기대지수": int(k.get("expectedClickScore") or 0),
            "평균 CPC": int(round(float(s.get("cpc", 0) or 0))),
            "총비용": int(round(float(s.get("salesAmt", 0) or 0))),
        })
    return pd.DataFrame(rows)


# ========================================================
#  채널 현황 (유튜브 구독자)
# ========================================================
YT_CHANNEL_ID = "UCbofDz8L4pyoBEjOEmM24GA"   # 코라스로보틱스


@st.cache_data(ttl=3600)
def yt_stats():
    key = st.secrets.get("YOUTUBE_API_KEY")
    if not key:
        return None
    try:
        url = "https://www.googleapis.com/youtube/v3/channels?" + urllib.parse.urlencode(
            {"part": "snippet,statistics", "id": YT_CHANNEL_ID, "key": key})
        with urllib.request.urlopen(url, timeout=30) as r:
            data = json.loads(r.read().decode())
        items = data.get("items", [])
        if not items:
            return None
        s = items[0]["statistics"]
        return {"title": items[0]["snippet"]["title"],
                "subs": int(s.get("subscriberCount", 0)),
                "views": int(s.get("viewCount", 0)),
                "videos": int(s.get("videoCount", 0))}
    except Exception:
        return None


@st.cache_data(ttl=3600)
def yt_history():
    """youtube_subs.csv(시드+누적) 를 읽고, 오늘 값이 없으면 실시간 구독자수를 끝점으로 덧붙임."""
    try:
        h = pd.read_csv("youtube_subs.csv")
        h["date"] = pd.to_datetime(h["date"])
    except Exception:
        h = pd.DataFrame(columns=["date", "subscribers"])
    today = pd.Timestamp(datetime.date.today())
    live = yt_stats()
    if live and (h.empty or today not in set(h["date"])):
        h = pd.concat([h, pd.DataFrame([{"date": today, "subscribers": live["subs"]}])],
                      ignore_index=True)
    h = h.dropna().sort_values("date")
    h["subscribers"] = h["subscribers"].astype(int)
    return h


# ========================================================
#  공통: 사이드바 (페이지 + 기간)
# ========================================================
df = load_sns()
if df.empty:
    st.warning("SNS/DA 데이터가 없습니다.")
    st.stop()
min_d = df["date"].min().date()
data_max = df["date"].max().date()
today_d = datetime.date.today()
# 달력 최대 선택일 = 오늘 (데이터 마지막 날이 아니라). 날짜 바뀌면 자동으로 따라감.
max_d = max(data_max, today_d)

st.sidebar.title("Koras 광고")
page = st.sidebar.radio("페이지", ["🏠 총계 (주간 보고)", "📊 유튜브 · 메타", "🔍 네이버 검색광고", "🔎 구글 검색광고", "📺 채널 현황"])
st.sidebar.divider()

# 기본 기간 = 이번 달 1일 ~ 오늘
default_start = today_d.replace(day=1)
if default_start < min_d:
    default_start = min_d
date_range = st.sidebar.date_input("기간", value=(default_start, max_d),
                                   min_value=min_d, max_value=max_d)
if isinstance(date_range, (list, tuple)) and len(date_range) == 2:
    start, end = date_range
else:
    start = end = date_range if not isinstance(date_range, (list, tuple)) else date_range[0]

period_len = (end - start).days + 1
prev_end = start - timedelta(days=1)
prev_start = prev_end - timedelta(days=period_len - 1)

if st.sidebar.button("🔄 데이터 새로고침"):
    st.cache_data.clear()
    st.rerun()
st.sidebar.caption(f"증감 = 선택 기간 vs 직전 {period_len}일 대비")


def pill(cur_v, prev_v):
    if not prev_v:
        return '<span class="k-pill k-up" style="opacity:0.7;">신규</span>'
    pct = (cur_v - prev_v) / prev_v * 100
    if pct >= 0:
        return f'<span class="k-pill k-up">▲ {pct:.1f}%</span>'
    return f'<span class="k-pill k-dn">▼ {abs(pct):.1f}%</span>'


def pilli(cur_v, prev_v):
    """히어로 하단 줄용 인라인 증감 뱃지 (위 칸이랑 같은 초록▲/빨강▼)."""
    if not prev_v:
        return '<span class="k-pill k-up" style="margin-top:0;">신규</span>'
    pct = (cur_v - prev_v) / prev_v * 100
    if pct >= 0:
        return f'<span class="k-pill k-up" style="margin-top:0;">▲ {pct:.1f}%</span>'
    return f'<span class="k-pill k-dn" style="margin-top:0;">▼ {abs(pct):.1f}%</span>'


def period_txt():
    return f"{start.strftime('%Y.%m.%d')} – {end.strftime('%m.%d')}"


# ========================================================
#  엑셀 보고서 (시트 3개: SNS·DA / 검색광고 / 채널현황)
# ========================================================
def _sns_table(d):
    """플랫폼별 상세 표 DataFrame (구글/메타/총계)."""
    rows = []
    for name, sub in [("구글", d[d["platform"] == "google"]),
                      ("메타", d[d["platform"] == "meta"]),
                      ("총계", d)]:
        imp = int(sub["impressions"].sum()); clk = int(sub["clicks"].sum())
        cost = float(sub["cost"].sum())
        ctr = (clk / imp * 100) if imp else 0
        cpc = (cost / clk) if clk else 0
        rows.append({"구분": name, "조회·도달": int(sub["views"].sum()),
                     "클릭": clk, "전환": int(round(sub["conversions"].sum())),
                     "노출": imp, "비용(원)": round(cost),
                     "CTR(%)": round(ctr, 2), "CPC(원)": round(cpc)})
    return pd.DataFrame(rows)


def build_excel_report():
    import io
    buf = io.BytesIO()
    sns_only = df[df["platform"].isin(["google", "meta"])]
    d = sns_only[(sns_only["date"].dt.date >= start) & (sns_only["date"].dt.date <= end)].copy()

    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        # ---- 시트 1: SNS·DA ----
        _sns_table(d).to_excel(writer, sheet_name="SNS·DA", index=False, startrow=1)
        daily_sns = (d.groupby([d["date"].dt.date, "platform"])
                     [["views", "clicks", "conversions", "impressions", "cost"]]
                     .sum().reset_index())
        daily_sns.columns = ["날짜", "플랫폼", "조회·도달", "클릭", "전환", "노출", "비용(원)"]
        daily_sns.to_excel(writer, sheet_name="SNS·DA", index=False,
                           startrow=len(_sns_table(d)) + 5)

        # ---- 시트 2: 검색광고 (네이버) ----
        if _naver_ready():
            try:
                kw = nv_keywords(start.isoformat(), end.isoformat())
                nd = nv_daily(start.isoformat(), end.isoformat())
                kw.sort_values("노출수", ascending=False).to_excel(
                    writer, sheet_name="검색광고(네이버)", index=False, startrow=1)
                if not nd.empty:
                    nd2 = nd.copy()
                    nd2["date"] = nd2["date"].dt.date
                    nd2.columns = ["날짜", "노출", "클릭", "비용(원)"]
                    nd2.to_excel(writer, sheet_name="검색광고(네이버)", index=False,
                                 startrow=len(kw) + 5)
            except Exception:
                pd.DataFrame({"안내": ["네이버 데이터를 불러오지 못했어요."]}).to_excel(
                    writer, sheet_name="검색광고(네이버)", index=False)
        else:
            pd.DataFrame({"안내": ["네이버 키가 설정되지 않았어요."]}).to_excel(
                writer, sheet_name="검색광고(네이버)", index=False)

        # ---- 시트 3: 채널현황 (유튜브 구독자) ----
        try:
            h = yt_history()
            hp = h[(h["date"].dt.date >= start) & (h["date"].dt.date <= end)].copy()
            hp["date"] = hp["date"].dt.date
            hp.columns = ["날짜", "구독자수"]
            hp.to_excel(writer, sheet_name="채널현황(유튜브)", index=False, startrow=1)
        except Exception:
            pd.DataFrame({"안내": ["유튜브 데이터를 불러오지 못했어요."]}).to_excel(
                writer, sheet_name="채널현황(유튜브)", index=False)

        # ---- 가벼운 서식: 제목 행 + 열 너비 ----
        wb = writer.book
        from openpyxl.styles import Font
        titles = {"SNS·DA": f"SNS·DA 성과  ({start} ~ {end})",
                  "검색광고(네이버)": f"검색광고(네이버) 키워드 성과  ({start} ~ {end})",
                  "채널현황(유튜브)": f"유튜브 구독자 추이  ({start} ~ {end})"}
        for name, ws in ((n, wb[n]) for n in wb.sheetnames):
            if name in titles:
                ws["A1"] = titles[name]
                ws["A1"].font = Font(bold=True, size=13, color="2563EB")
            for col in ws.columns:
                width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
                ws.column_dimensions[col[0].column_letter].width = min(max(width + 4, 10), 30)

    buf.seek(0)
    return buf.getvalue()


st.sidebar.divider()
if st.sidebar.button("📥 엑셀 보고서 생성"):
    with st.spinner("보고서 만드는 중…"):
        st.session_state["xlsx_report"] = build_excel_report()
if "xlsx_report" in st.session_state:
    st.sidebar.download_button(
        "⬇️ 엑셀 다운로드",
        st.session_state["xlsx_report"],
        file_name=f"koras_ad_report_{start}_{end}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    st.sidebar.caption("선택한 기간 기준 · 시트 3개 (SNS·DA / 검색광고 / 채널현황)")


# ========================================================
#  페이지 1 — SNS / DA
# ========================================================
if page.startswith("📊"):
    sns_df = df[df["platform"].isin(["google", "meta"])]
    f = sns_df[(sns_df["date"].dt.date >= start) & (sns_df["date"].dt.date <= end)].copy()
    f_prev = sns_df[(sns_df["date"].dt.date >= prev_start) & (sns_df["date"].dt.date <= prev_end)].copy()

    def agg(d):
        return {"views": int(d["views"].sum()), "clicks": int(d["clicks"].sum()),
                "conversions": int(round(d["conversions"].sum())),
                "impressions": int(d["impressions"].sum()), "cost": float(d["cost"].sum())}

    cur = agg(f); prev = agg(f_prev)

    st.markdown(f"""
    <div style="display:flex; align-items:flex-end; justify-content:space-between; gap:12px; margin-bottom:16px;">
      <div><div class="k-tag">KORAS ROBOTICS · 광고 리포트</div><div class="k-h1">SNS · DA 통합 성과</div></div>
      <div style="text-align:right; font-size:12px; color:rgba(128,128,128,0.95); line-height:1.6;">
        <div>{period_txt()}</div><div style="opacity:0.7;">직전 {period_len}일 대비 · 구글 + 메타</div></div>
    </div>""", unsafe_allow_html=True)

    t_ctr = (cur["clicks"] / cur["impressions"] * 100) if cur["impressions"] else 0
    t_cpc = (cur["cost"] / cur["clicks"]) if cur["clicks"] else 0
    p_ctr = (prev["clicks"] / prev["impressions"] * 100) if prev["impressions"] else 0
    p_cpc = (prev["cost"] / prev["clicks"]) if prev["clicks"] else 0
    st.markdown(f"""
    <div class="k-hero"><div style="display:grid; grid-template-columns:repeat(4,1fr); gap:10px;">
      <div class="k-cell"><div class="lbl">조회·도달</div><div class="num">{cur['views']:,}</div>{pill(cur['views'], prev['views'])}</div>
      <div class="k-cell"><div class="lbl">클릭수</div><div class="num">{cur['clicks']:,}</div>{pill(cur['clicks'], prev['clicks'])}</div>
      <div class="k-cell"><div class="lbl">전환수</div><div class="num">{cur['conversions']:,}</div>{pill(cur['conversions'], prev['conversions'])}</div>
      <div class="k-cell"><div class="lbl">노출수</div><div class="num">{cur['impressions']:,}</div>{pill(cur['impressions'], prev['impressions'])}</div>
    </div><div class="k-strip"><span>비용 <b>{cur['cost']:,.0f}원</b> {pilli(cur['cost'], prev['cost'])}</span><span>CTR <b>{t_ctr:.2f}%</b> {pilli(t_ctr, p_ctr)}</span><span>CPC <b>{t_cpc:,.0f}원</b> {pilli(t_cpc, p_cpc)}</span></div></div>
    """, unsafe_allow_html=True)
    st.markdown('<div style="font-size:11px; color:rgba(128,128,128,0.8); margin:6px 0 18px;">※ \'조회·도달\'은 구글 조회수 + 메타 도달의 합이에요(성격이 다른 값이라 참고용).</div>', unsafe_allow_html=True)

    # 표
    def cells(d):
        a = agg(d)
        ctr = (a["clicks"] / a["impressions"] * 100) if a["impressions"] else 0
        cpc = (a["cost"] / a["clicks"]) if a["clicks"] else 0
        return [f"{a['views']:,}", f"{a['clicks']:,}", f"{a['conversions']:,}",
                f"{a['impressions']:,}", f"{a['cost']:,.0f}원", f"{ctr:.2f}%", f"{cpc:,.0f}원"]

    st.markdown('<div class="k-sec"><span class="k-bar"></span><span class="k-sec-t">플랫폼별 상세</span></div>', unsafe_allow_html=True)
    hd = ["구분", "조회·도달", "클릭", "전환", "노출", "비용", "CTR", "CPC"]
    rdef = [("구글", "#2563EB", cells(f[f["platform"] == "google"]), False),
            ("메타", "#7AA5F5", cells(f[f["platform"] == "meta"]), False),
            ("총계", None, cells(f), True)]
    html = '<div class="k-wrap"><table class="k-tbl"><thead><tr>'
    for i, h in enumerate(hd):
        html += f"<th style='text-align:{'left' if i == 0 else 'right'};'>{h}</th>"
    html += "</tr></thead><tbody>"
    for name, color, vals, tot in rdef:
        cls = " class='tot'" if tot else ""
        dot = f"<span class='k-dot' style='background:{color};'></span>" if color else ""
        html += f"<tr{cls}><td style='text-align:left;'>{dot}{name}</td>"
        for v in vals:
            html += f"<td style='text-align:right;'>{v}</td>"
        html += "</tr>"
    html += "</tbody></table></div>"
    st.markdown(html, unsafe_allow_html=True)
    st.markdown("<div style='height:18px;'></div>", unsafe_allow_html=True)

    # 그래프
    LABELS = {"views": "조회·도달", "clicks": "클릭수", "conversions": "전환수", "impressions": "노출수"}
    pick = st.radio("플랫폼", ["전체", "구글", "메타"], horizontal=True)
    if pick == "구글":
        g = f[f["platform"] == "google"]; unit = "캠페인별"
    elif pick == "메타":
        g = f[f["platform"] == "meta"]; unit = "광고세트별"
    else:
        g = f; unit = "캠페인/광고세트별"

    if g.empty:
        st.info("선택한 조건에 데이터가 없어요.")
    else:
        st.markdown('<div class="k-sec" style="margin-top:6px;"><span class="k-bar"></span><span class="k-sec-t">일별 추이</span></div>', unsafe_allow_html=True)
        mk = ["views", "clicks", "conversions", "impressions"]; ml = [LABELS[k] for k in mk]
        chosen = st.multiselect("표시할 지표", ml, default=ml)
        if chosen:
            cols = [k for k in mk if LABELS[k] in chosen]
            daily = g.groupby("date")[cols].sum().reset_index()
            long = daily.melt("date", var_name="m", value_name="값")
            long["지표"] = long["m"].map(LABELS)
            long["상대값"] = long.groupby("m")["값"].transform(lambda s: s / s.max() * 100 if s.max() else s * 0)
            st.altair_chart(alt.Chart(long).mark_line(point=True, strokeWidth=2.5).encode(
                x=alt.X("date:T", title="날짜"),
                y=alt.Y("상대값:Q", title="상대값 (지표별 최대=100)"),
                color=alt.Color("지표:N", title="지표", scale=alt.Scale(range=BLUE_SCHEME)),
                tooltip=["date:T", "지표:N", alt.Tooltip("값:Q", title="실제값", format=",.0f")],
            ).properties(height=360), width="stretch")
            st.caption("※ 지표마다 단위가 달라, 각 지표를 '자기 최대값=100' 기준으로 맞춰 그렸어요.")

        st.markdown(f'<div class="k-sec" style="margin-top:8px;"><span class="k-bar"></span><span class="k-sec-t">{unit} 비교</span></div>', unsafe_allow_html=True)
        bl = st.selectbox("지표 선택", ml, index=0)
        bcol = [k for k in mk if LABELS[k] == bl][0]
        bar_df = g.groupby("campaign")[bcol].sum().reset_index().sort_values(bcol, ascending=False)
        st.altair_chart(alt.Chart(bar_df).mark_bar(color=BLUE, cornerRadiusTopLeft=4, cornerRadiusTopRight=4).encode(
            x=alt.X("campaign:N", sort="-y", title=None, axis=alt.Axis(labelAngle=-35)),
            y=alt.Y(f"{bcol}:Q", title=bl),
            tooltip=[alt.Tooltip("campaign:N", title="이름"), alt.Tooltip(f"{bcol}:Q", title=bl, format=",.0f")],
        ).properties(height=380), width="stretch")


# ========================================================
#  페이지 2 — 검색광고 (네이버)
# ========================================================
elif page.startswith("🔍"):
    st.markdown(f"""
    <div style="display:flex; align-items:flex-end; justify-content:space-between; gap:12px; margin-bottom:16px;">
      <div><div class="k-tag">KORAS ROBOTICS · 검색광고</div><div class="k-h1">검색광고(SA) 성과 · 네이버</div></div>
      <div style="text-align:right; font-size:12px; color:rgba(128,128,128,0.95); line-height:1.6;">
        <div>{period_txt()}</div><div style="opacity:0.7;">직전 {period_len}일 대비 · 파워링크</div></div>
    </div>""", unsafe_allow_html=True)

    if not _naver_ready():
        st.info("네이버 키가 설정되지 않았어요. Streamlit Secrets에 NAVER_CUSTOMER_ID / NAVER_API_KEY / NAVER_SECRET_KEY 를 추가하면 이 페이지가 켜져요.")
        st.stop()

    since, until = start.isoformat(), end.isoformat()
    p_since, p_until = prev_start.isoformat(), prev_end.isoformat()

    with st.spinner("네이버 데이터를 불러오는 중…"):
        cur = nv_summary(since, until)
        prev = nv_summary(p_since, p_until)
        daily = nv_daily(since, until)
        kw = nv_keywords(since, until)

    t_ctr = (cur["clicks"] / cur["impressions"] * 100) if cur["impressions"] else 0
    t_cpc = (cur["cost"] / cur["clicks"]) if cur["clicks"] else 0
    p_ctr = (prev["clicks"] / prev["impressions"] * 100) if prev["impressions"] else 0
    p_cpc = (prev["cost"] / prev["clicks"]) if prev["clicks"] else 0
    st.markdown(f"""
    <div class="k-hero"><div style="display:grid; grid-template-columns:repeat(4,1fr); gap:10px;">
      <div class="k-cell"><div class="lbl">노출수</div><div class="num">{cur['impressions']:,}</div>{pill(cur['impressions'], prev['impressions'])}</div>
      <div class="k-cell"><div class="lbl">클릭수</div><div class="num">{cur['clicks']:,}</div>{pill(cur['clicks'], prev['clicks'])}</div>
      <div class="k-cell"><div class="lbl">전환수</div><div class="num">{int(cur['conversions']):,}</div>{pill(cur['conversions'], prev['conversions'])}</div>
      <div class="k-cell"><div class="lbl">평균 CPC</div><div class="num">{t_cpc:,.0f}<span style="font-size:15px;">원</span></div>{pill(t_cpc, p_cpc)}</div>
    </div><div class="k-strip"><span>CTR <b>{t_ctr:.2f}%</b> {pilli(t_ctr, p_ctr)}</span><span>총비용 <b>{cur['cost']:,.0f}원</b> {pilli(cur['cost'], prev['cost'])}</span></div></div>
    """, unsafe_allow_html=True)
    st.markdown("<div style='height:18px;'></div>", unsafe_allow_html=True)

    # 차트 (노출수 / 클릭수)
    st.markdown('<div class="k-sec"><span class="k-bar"></span><span class="k-sec-t">일별 추이 (노출수 · 클릭수)</span></div>', unsafe_allow_html=True)
    if daily.empty or daily["impressions"].sum() == 0:
        st.info("선택 기간에 네이버 실적이 없어요.")
    else:
        NL = {"impressions": "노출수", "clicks": "클릭수"}
        long = daily.melt("date", value_vars=["impressions", "clicks"], var_name="m", value_name="값")
        long["지표"] = long["m"].map(NL)
        long["상대값"] = long.groupby("m")["값"].transform(lambda s: s / s.max() * 100 if s.max() else s * 0)
        st.altair_chart(alt.Chart(long).mark_line(point=True, strokeWidth=2.5).encode(
            x=alt.X("date:T", title="날짜"),
            y=alt.Y("상대값:Q", title="상대값 (지표별 최대=100)"),
            color=alt.Color("지표:N", title="지표", scale=alt.Scale(range=[BLUE, "#7AA5F5"])),
            tooltip=["date:T", "지표:N", alt.Tooltip("값:Q", title="실제값", format=",.0f")],
        ).properties(height=340), width="stretch")
        st.caption("※ 노출수와 클릭수는 단위 차이가 커서 각각 '최대=100' 기준으로 맞춰 그렸어요. 선에 마우스를 올리면 실제 숫자가 나와요.")

    # 키워드 표 (헤더 클릭으로 오름/내림 정렬)
    st.markdown("<div style='height:8px;'></div>", unsafe_allow_html=True)
    st.markdown('<div class="k-sec"><span class="k-bar"></span><span class="k-sec-t">키워드별 상세</span></div>', unsafe_allow_html=True)
    st.caption("열 제목을 클릭하면 오름차순/내림차순으로 정렬돼요.")
    if kw.empty:
        st.info("키워드 데이터가 없어요.")
    else:
        kw_sorted = kw.sort_values("노출수", ascending=False)
        st.dataframe(
            kw_sorted, width="stretch", hide_index=True, height=560,
            column_config={
                "키워드": st.column_config.TextColumn("키워드", width="medium"),
                "노출수": st.column_config.NumberColumn("노출수", format="%d"),
                "클릭수": st.column_config.NumberColumn("클릭수", format="%d"),
                "현재 입찰가": st.column_config.NumberColumn("현재 입찰가", format="%d원"),
                "클릭 기대지수": st.column_config.NumberColumn("클릭 기대지수", format="%d/10"),
                "평균 CPC": st.column_config.NumberColumn("평균 CPC", format="%d원"),
                "총비용": st.column_config.NumberColumn("총비용", format="%d원"),
            },
        )


# ========================================================
#  페이지 3 — 채널 현황 (유튜브 구독자)
# ========================================================
elif page.startswith("📺"):
    st.markdown(f"""
    <div style="display:flex; align-items:flex-end; justify-content:space-between; gap:12px; margin-bottom:16px;">
      <div><div class="k-tag">KORAS ROBOTICS · 채널 현황</div><div class="k-h1">유튜브 채널 성장</div></div>
      <div style="text-align:right; font-size:12px; color:rgba(128,128,128,0.95); line-height:1.6;">
        <div>{period_txt()}</div><div style="opacity:0.7;">구독자 추이</div></div>
    </div>""", unsafe_allow_html=True)

    stats = yt_stats()
    hist = yt_history()

    if stats is None:
        st.info("유튜브 키가 설정되지 않았어요. Streamlit Secrets에 YOUTUBE_API_KEY 를 추가하면 이 페이지가 켜져요.")
        st.stop()

    # 선택 기간 내 구독자 증감 (기간 시작값 대비 현재)
    h_period = hist[(hist["date"].dt.date >= start) & (hist["date"].dt.date <= end)]
    cur_subs = stats["subs"]
    base_subs = int(h_period["subscribers"].iloc[0]) if not h_period.empty else cur_subs
    diff = cur_subs - base_subs

    def num_pill(diff):
        if diff > 0:
            return f'<span class="k-pill k-up" style="margin-top:0;">▲ {diff:,}명</span>'
        if diff < 0:
            return f'<span class="k-pill k-dn" style="margin-top:0;">▼ {abs(diff):,}명</span>'
        return '<span class="k-pill k-up" style="margin-top:0; opacity:0.7;">변동 없음</span>'

    st.markdown(f"""
    <div class="k-hero"><div style="display:grid; grid-template-columns:repeat(3,1fr); gap:10px;">
      <div class="k-cell"><div class="lbl">구독자 수</div><div class="num">{stats['subs']:,}</div>
        <div style="margin-top:9px;">{num_pill(diff)} <span style="font-size:11px; color:rgba(128,128,128,0.9);">기간 내</span></div></div>
      <div class="k-cell"><div class="lbl">총 조회수</div><div class="num">{stats['views']:,}</div></div>
      <div class="k-cell"><div class="lbl">영상 수</div><div class="num">{stats['videos']:,}</div></div>
    </div><div class="k-strip"><span>채널 <b>{stats['title']}</b></span><span>실시간 구독자 기준</span></div></div>
    """, unsafe_allow_html=True)
    st.markdown("<div style='height:18px;'></div>", unsafe_allow_html=True)

    # 구독자 추이 차트 (선택 기간)
    st.markdown('<div class="k-sec"><span class="k-bar"></span><span class="k-sec-t">구독자 추이</span></div>', unsafe_allow_html=True)
    if h_period.empty:
        st.info("선택 기간에 구독자 기록이 없어요. 기간을 넓혀보세요.")
    else:
        chart = (
            alt.Chart(h_period).mark_line(point=True, strokeWidth=2.5, color=BLUE).encode(
                x=alt.X("date:T", title="날짜", scale=alt.Scale(padding=20)),
                y=alt.Y("subscribers:Q", title="구독자 수", scale=alt.Scale(zero=False)),
                tooltip=[alt.Tooltip("date:T", title="날짜"),
                         alt.Tooltip("subscribers:Q", title="구독자", format=",.0f")],
            ).properties(height=380, padding={"right": 28, "left": 4, "top": 28, "bottom": 4})
        )
        st.altair_chart(chart, width="stretch")
        st.caption("※ 6/25까지는 입력해둔 기록, 6/26부터는 매일 자동으로 쌓여요. (현재 구독자 수는 실시간)")


# ========================================================
#  페이지 — 🔎 구글 검색광고 (SA)
# ========================================================
if page.startswith("🔎"):
    st.markdown(f"""
    <div style="display:flex; align-items:flex-end; justify-content:space-between; gap:12px; margin-bottom:16px;">
      <div><div class="k-tag">KORAS ROBOTICS · 검색광고</div><div class="k-h1">검색광고(SA) 성과 · 구글</div></div>
      <div style="text-align:right; font-size:12px; color:rgba(128,128,128,0.95); line-height:1.6;">
        <div>{period_txt()}</div><div style="opacity:0.7;">직전 {period_len}일 대비 · 구글 검색</div></div>
    </div>""", unsafe_allow_html=True)

    gsa = df[df["platform"] == "google_sa"]
    g = gsa[(gsa["date"].dt.date >= start) & (gsa["date"].dt.date <= end)].copy()
    g_prev = gsa[(gsa["date"].dt.date >= prev_start) & (gsa["date"].dt.date <= prev_end)].copy()

    if g.empty:
        st.info("선택 기간에 구글 검색광고 데이터가 없어요. (캠페인 집행 후 하루 이틀 지나면 잡혀요.)")
    else:
        def gagg(d):
            return {"clicks": int(d["clicks"].sum()),
                    "conversions": int(round(d["conversions"].sum())),
                    "impressions": int(d["impressions"].sum()),
                    "cost": float(d["cost"].sum())}
        cur = gagg(g); prev = gagg(g_prev)
        t_ctr = (cur["clicks"] / cur["impressions"] * 100) if cur["impressions"] else 0
        t_cpc = (cur["cost"] / cur["clicks"]) if cur["clicks"] else 0
        p_ctr = (prev["clicks"] / prev["impressions"] * 100) if prev["impressions"] else 0
        p_cpc = (prev["cost"] / prev["clicks"]) if prev["clicks"] else 0

        st.markdown(f"""
        <div class="k-hero"><div style="display:grid; grid-template-columns:repeat(4,1fr); gap:10px;">
          <div class="k-cell"><div class="lbl">노출수</div><div class="num">{cur['impressions']:,}</div>{pill(cur['impressions'], prev['impressions'])}</div>
          <div class="k-cell"><div class="lbl">클릭수</div><div class="num">{cur['clicks']:,}</div>{pill(cur['clicks'], prev['clicks'])}</div>
          <div class="k-cell"><div class="lbl">전환수</div><div class="num">{cur['conversions']:,}</div>{pill(cur['conversions'], prev['conversions'])}</div>
          <div class="k-cell"><div class="lbl">평균 CPC</div><div class="num">{t_cpc:,.0f}<span style="font-size:15px;">원</span></div>{pill(t_cpc, p_cpc)}</div>
        </div><div class="k-strip"><span>CTR <b>{t_ctr:.2f}%</b> {pilli(t_ctr, p_ctr)}</span><span>총비용 <b>{cur['cost']:,.0f}원</b> {pilli(cur['cost'], prev['cost'])}</span></div></div>
        """, unsafe_allow_html=True)
        st.markdown("<div style='height:18px;'></div>", unsafe_allow_html=True)

        st.markdown('<div class="k-sec"><span class="k-bar"></span><span class="k-sec-t">일별 추이 (노출수 · 클릭수)</span></div>', unsafe_allow_html=True)
        GL = {"impressions": "노출수", "clicks": "클릭수"}
        daily = g.groupby("date")[["impressions", "clicks"]].sum().reset_index()
        long = daily.melt("date", var_name="m", value_name="값")
        long["지표"] = long["m"].map(GL)
        long["상대값"] = long.groupby("m")["값"].transform(lambda s: s / s.max() * 100 if s.max() else s * 0)
        st.altair_chart(alt.Chart(long).mark_line(point=True, strokeWidth=2.5).encode(
            x=alt.X("date:T", title="날짜"),
            y=alt.Y("상대값:Q", title="상대값 (지표별 최대=100)"),
            color=alt.Color("지표:N", title="지표", scale=alt.Scale(range=[BLUE, "#7AA5F5"])),
            tooltip=["date:T", "지표:N", alt.Tooltip("값:Q", title="실제값", format=",.0f")],
        ).properties(height=340), width="stretch")

        st.markdown('<div class="k-sec" style="margin-top:8px;"><span class="k-bar"></span><span class="k-sec-t">캠페인별 비교</span></div>', unsafe_allow_html=True)
        gl_labels = ["노출수", "클릭수", "전환수"]
        gl_keys = {"노출수": "impressions", "클릭수": "clicks", "전환수": "conversions"}
        bl = st.selectbox("지표 선택", gl_labels, index=0, key="gsa_bar")
        bcol = gl_keys[bl]
        bar_df = g.groupby("campaign")[bcol].sum().reset_index().sort_values(bcol, ascending=False)
        st.altair_chart(alt.Chart(bar_df).mark_bar(color=BLUE, cornerRadiusTopLeft=4, cornerRadiusTopRight=4).encode(
            x=alt.X("campaign:N", sort="-y", title=None, axis=alt.Axis(labelAngle=-35)),
            y=alt.Y(f"{bcol}:Q", title=bl),
            tooltip=[alt.Tooltip("campaign:N", title="캠페인"), alt.Tooltip(f"{bcol}:Q", title=bl, format=",.0f")],
        ).properties(height=380), width="stretch")


# ========================================================
#  페이지 — 🏠 총계 (주간 보고, 메인)
# ========================================================
elif page.startswith("🏠"):
    today_dt = datetime.date.today()
    week_start = today_dt - timedelta(days=6)                       # 최근 7일 (오늘 포함)
    pw_end = week_start - timedelta(days=1)                         # 직전 7일
    pw_start = pw_end - timedelta(days=6)

    st.markdown(f"""
    <div style="display:flex; align-items:flex-end; justify-content:space-between; gap:12px; margin-bottom:16px;">
      <div><div class="k-tag">KORAS ROBOTICS · 주간 보고</div><div class="k-h1">이번 주 성과 한눈에</div></div>
      <div style="text-align:right; font-size:12px; color:rgba(128,128,128,0.95); line-height:1.6;">
        <div>{week_start.strftime('%m.%d')} – {today_dt.strftime('%m.%d')} (최근 7일)</div>
        <div style="opacity:0.7;">직전 7일({pw_start.strftime('%m.%d')}–{pw_end.strftime('%m.%d')}) 대비</div></div>
    </div>""", unsafe_allow_html=True)

    # ---- 이번 주 / 지난주 데이터 ----
    def week_slice(d, s, e):
        return d[(d["date"].dt.date >= s) & (d["date"].dt.date <= e)]

    cur_all = week_slice(df, week_start, today_dt)
    prev_all = week_slice(df, pw_start, pw_end)

    # 노출 총계 (유튜브+메타+구글SA + 네이버)
    cur_imp = int(cur_all["impressions"].sum())
    prev_imp = int(prev_all["impressions"].sum())
    if _naver_ready():
        try:
            cur_imp += int(nv_summary(week_start.isoformat(), today_dt.isoformat())["impressions"])
            prev_imp += int(nv_summary(pw_start.isoformat(), pw_end.isoformat())["impressions"])
        except Exception:
            pass

    # 광고로 만난 사람 (구글 조회수 + 메타 도달)
    cur_reach = int(cur_all[cur_all["platform"].isin(["google", "meta"])]["views"].sum())
    prev_reach = int(prev_all[prev_all["platform"].isin(["google", "meta"])]["views"].sum())

    # 유튜브 구독자 / 총 조회수
    stats = yt_stats()
    hist = yt_history()
    subs_now = stats["subs"] if stats else 0
    total_views = stats["views"] if stats else 0
    hw = hist[(hist["date"].dt.date >= week_start) & (hist["date"].dt.date <= today_dt)]
    subs_week_gain = (subs_now - int(hw["subscribers"].iloc[0])) if not hw.empty else 0

    def gain_pill(diff, unit="명"):
        if diff > 0:
            return f'<span class="k-pill k-up">▲ {diff:,}{unit} 증가</span>'
        if diff < 0:
            return f'<span class="k-pill k-dn">▼ {abs(diff):,}{unit}</span>'
        return '<span class="k-pill k-up" style="opacity:0.7;">변동 없음</span>'

    # ---- 상단: 노출 총계 + 유튜브 구독자 + 총 조회수 ----
    st.markdown(f"""
    <div class="k-hero"><div style="display:grid; grid-template-columns:repeat(3,1fr); gap:10px;">
      <div class="k-cell"><div class="lbl">최근 7일, 회사가 노출된 횟수</div><div class="num">{cur_imp:,}</div>{pill(cur_imp, prev_imp)}</div>
      <div class="k-cell"><div class="lbl">유튜브 구독자</div><div class="num">{subs_now:,}</div>{gain_pill(subs_week_gain)}</div>
      <div class="k-cell"><div class="lbl">유튜브 총 조회수 (누적)</div><div class="num">{total_views:,}</div></div>
    </div><div class="k-strip"><span>유튜브 · 메타 · 네이버 · 구글 검색광고 합산</span></div></div>
    """, unsafe_allow_html=True)
    st.markdown("<div style='height:18px;'></div>", unsafe_allow_html=True)

    # ---- 구독자 성장 곡선 (크게, 전체 기간) ----
    st.markdown('<div class="k-sec"><span class="k-bar"></span><span class="k-sec-t">유튜브 구독자 성장</span></div>', unsafe_allow_html=True)
    if hist.empty:
        st.info("구독자 기록이 아직 없어요.")
    else:
        st.altair_chart(alt.Chart(hist).mark_area(
            line={"color": BLUE, "strokeWidth": 3},
            color=alt.Gradient(gradient="linear",
                               stops=[alt.GradientStop(color="#DBEAFE", offset=0),
                                      alt.GradientStop(color="#2563EB", offset=1)],
                               x1=1, x2=1, y1=1, y2=0),
        ).encode(
            x=alt.X("date:T", title=None, scale=alt.Scale(padding=20)),
            y=alt.Y("subscribers:Q", title="구독자 수", scale=alt.Scale(zero=False)),
            tooltip=[alt.Tooltip("date:T", title="날짜"),
                     alt.Tooltip("subscribers:Q", title="구독자", format=",.0f")],
        ).properties(height=340, padding={"right": 28, "top": 28}), width="stretch")

    # ---- 하단: 광고로 만난 사람 (주간) ----
    st.markdown('<div class="k-sec" style="margin-top:6px;"><span class="k-bar"></span><span class="k-sec-t">최근 7일, 광고로 만난 사람</span></div>', unsafe_allow_html=True)
    c1, c2 = st.columns([1, 2])
    with c1:
        st.markdown(f"""
        <div class="k-hero" style="margin-bottom:0;"><div class="k-cell">
          <div class="lbl">광고를 본 사람 수</div>
          <div class="num">{cur_reach:,}</div>
          {pill(cur_reach, prev_reach)}
        </div></div>""", unsafe_allow_html=True)
        st.caption("유튜브 영상 조회 + 메타(인스타·페이스북) 도달 합산 · 직전 7일 대비")
    with c2:
        wk = cur_all[cur_all["platform"].isin(["google", "meta"])]
        if not wk.empty:
            dv = wk.groupby("date")["views"].sum().reset_index()
            st.altair_chart(alt.Chart(dv).mark_bar(color=BLUE, cornerRadiusTopLeft=4, cornerRadiusTopRight=4).encode(
                x=alt.X("date:T", title=None),
                y=alt.Y("views:Q", title="만난 사람"),
                tooltip=[alt.Tooltip("date:T", title="날짜"), alt.Tooltip("views:Q", title="만난 사람", format=",.0f")],
            ).properties(height=220), width="stretch")
